from selenium import webdriver

import time
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
import re
from webdriver_manager.chrome import ChromeDriverManager

#엑셀 파일 열기
wb = openpyxl.load_workbook('쿠팡.xlsx')
ws = wb.worksheets[0]
count = 2
c = 2
wb2 = openpyxl.load_workbook('order.xlsx')
ws2 = wb2.worksheets[0]

wb3 = openpyxl.load_workbook('GPS product.xlsx')
ws3 = wb3.worksheets[0]
productlib = {}
p = 2
while ws3["A"+str(p)].value != None:
    productlib[ws3["B"+str(p)].value] = [ws3["E"+str(p)].value, ws3["F"+str(p)].value, ws3["G"+str(p)].value, ws3["C"+str(p)].value, ws3["D"+str(p)].value, ws3["H"+str(p)].value]
    p = p+1

mylib = {}
z = 1
for row_cells in ws.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        cc = str(cell)
        if cell.value in mylib:
            pass
        else:
            if z == 1:
                mylib[cell.value] = cc[len(cc)-3:len(cc)-2]
            elif z != 1:
                mylib[cell.value] = cc[len(cc)-4:len(cc)-2]
            if mylib[cell.value] == "Z":
                z = 2
        
 

rdict = {}
rdictcheck = {}
rdicterror = {}
cccccc = 2
wb5 = openpyxl.load_workbook('ordererror.xlsx')
ws5 = wb5.worksheets[0]
while ws["A"+str(count)].value != None:
    rnumber = ws[mylib["Personal Customs Clearance Code (PCCC)"]+str(count)].value
    rname = ws[mylib["Buyer"]+str(count)].value
    rorder = ws[mylib["Order number"]+str(count)].value
    rproduct = ws[mylib["Registered product name"]+str(count)].value
    rdict[rnumber] = []
    if rnumber == None or rnumber == "":
        rdicterror[rname] = [rname, rorder, rproduct, rnumber]
    count = count+1

for i in rdicterror:
    ws5["A"+str(cccccc)].value = rdicterror[i][1]
    ws5["B"+str(cccccc)].value = rdicterror[i][0]
    ws5["C"+str(cccccc)].value = rdicterror[i][3]
    ws5["D"+str(cccccc)].value = rdicterror[i][2]
    cccccc=cccccc+1
wb5.save('ordererror1.xlsx')
if cccccc >=3:
    print("************ordererror1 파일 확인 부탁 드립니다!!!!!!! 송장번호 겹치는 사람 발견!!!!!!!!!!!!!!!!!!*************************************************")


op = 0
count = 2
while ws["A"+str(count)].value != None:
    buyer = ws[mylib["Buyer"]+str(count)].value 
    receiver = ws[mylib["Recipient name"]+str(count)].value
    if buyer == receiver:
        receiver = buyer
    elif buyer != receiver:
        receiver = buyer+"("+receiver+")"
    rnumber = ws[mylib["Personal Customs Clearance Code (PCCC)"]+str(count)].value
    if "1234" in str(rnumber):
        print(receiver + " 의 통관부호가 제대로 입력되지 않았습니다.")
    rphone = ws[mylib["Buyer phone number"]+str(count)].value
    rphone1 = ws[mylib["Contact information of buyer for customs clearance purpose"]+str(count)].value
    if "0000" in str(rphone):
        print(receiver + " 의 전화번호가 제대로 입력되지 않았습니다.")
    elif "0000" in str(rphone1):
        print(receiver + " 의 전화번호가 제대로 입력되지 않았습니다.")
    zipcode = ws[mylib["Zipcode"]+str(count)].value
    if "-" in zipcode:
        zipcode = zipcode.replace("-", "")
    address1 = ws[mylib["Recipient address"]+str(count)].value
    note = ws[mylib["Delivery message"]+str(count)].value
    pname = ws[mylib["Registered product name"]+str(count)].value
    pname = pname.replace(u'\ufeff', '')
    poption = ""
    p=pname
    if mylib.get("Displayed product name") != None:
        poption = ws[mylib["Displayed product name"]+str(count)].value
        if poption != None:
            if pname in productlib:
                p = pname+str(op)
                s= productlib[pname][0]
                s1 = productlib[pname][1]
                s2 = productlib[pname][2]
                s3 = productlib[pname][3]
                s4 = productlib[pname][4]
                productlib[p] = [s+poption,s1,s2,s3,s4]
                op = op+1
            else:
                productlib[pname][0] = productlib[pname][0] + poption
    pnumber = ws[mylib["Order number"]+str(count)].value
    pamount = ws[mylib["Purchased qty"]+str(count)].value
    if productlib[pname][5] != None:
        pamount = int(ws[mylib["Purchased qty"]+str(count)].value) * int(productlib[pname][5])
    rdict[rnumber].append([receiver,rnumber,rphone, rphone1, zipcode,address1, note,p,pnumber,pamount])
    count = count + 1

    
hangul = re.compile('[^ ㄱ-ㅣ가-힣]+')
ff = sorted(rdict.items(), key=lambda e: e[1][0][7])


for x in ff:
    i = x[0]
    ws2["I"+str(c)].value = i
    for j in rdict[i]:
        ws2["C"+str(c)].value = j[0]  
        ws2["F"+str(c)].value = j[8]
        ws2["J"+str(c)].value = j[2]
        ws2["K"+str(c)].value = j[3]
        ws2["L"+str(c)].value = j[4]
        ws2["M"+str(c)].value = j[5]
        ws2["O"+str(c)].value = j[6]
        sss = productlib[j[7]][0]
        result = hangul.sub('', sss)
        result = hangul.findall(sss)
        ss = ""
        for v in result:
          ss = ss + str(v)+ " "
        ws2["P"+str(c)].value = ss
        ws2["Q"+str(c)].value = j[9]
        ws2["R"+str(c)].value = productlib[j[7]][1]
        ws2["G"+str(c)].value = productlib[j[7]][2]
        ws2["D"+str(c)].value = float(productlib[j[7]][3])
        ws2["S"+str(c)].value = productlib[j[7]][4]
        c= c+1

        

        
wb2.save('order1.xlsx')